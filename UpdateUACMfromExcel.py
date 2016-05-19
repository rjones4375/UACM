# from openpyxl import Workbook
from openpyxl.compat import range
# from openpyxl.cell import get_column_letter
import openpyxl as xl
import os.path
import pyodbc
import pprint

qrystr= ''
resultset = ''

conn_str = "DRIVER={SQL Server};SERVER=ed-cobalt.gsu.edu\sql2005;DATABASE=pact;UID=sa;PWD=whatisit"
connx = pyodbc.connect(conn_str)
qrystr = """
select Top 10 * from dbo.UACM_Candidate_Contact
"""
resultset = connx.execute(qrystr) 
ndx = 0
print ( resultset.description)
for row in resultset:
    print( 'Row {:d} {:d}'.format(ndx, row[1]))
    ndx = ndx+1

print('Opening workbook...')
   
wb = xl.load_workbook('RJ_Survey0ct2013.xlsx')
print( wb.sheetnames )

sheet = wb.get_sheet_by_name('Export')


colHeads = []
list_of_rows = {}

   # TODO: Fill in countyData with each county's population and tracts.
print('Reading rows...')
print( "Max Rows: %d  Max Cols: %d" % (sheet.max_row, sheet.max_column) )
# print( sheet.cell(row=1,column=1).value)

# read first row ( Column Headers )
sheetrow = sheet.rows[0]

# Each row in the spreadsheet has data for one survey.
for cellObj in sheetrow:
        colHeads.append(cellObj.value)
 
# for sheetrow in rows 2 to max_rows -- READ rest of sheet
ndx = 1
for ndx in range(1, sheet.max_row):
    thisrow = sheet.rows[ndx]
    # print( hex(id(thisrow)))
    rowData = {}
    j=0
    for cellObj in thisrow :
        rowData[colHeads[j]]= cellObj.value
        j += 1   
    list_of_rows[ndx] = rowData  # add this dictionary to 'list_of_rows' Dictionary

for i in range(1, 30):
     print( list_of_rows[i]['LastName'], 
           list_of_rows[i]['FirstName'], 
           list_of_rows[i]['PantherID'], 
           list_of_rows[i]['EmailAddress']
           )

     qrystr = """select * from dbo.UACM_Candidate_Contact
        where fk_PantherID = ?""" 
    
     resultset = connx.execute(qrystr, list_of_rows[i]['PantherID'] ) 
     ndx = 0
    # print ( resultset.description)
     for row in resultset:
        #print( row)
        print( '   ID KEY FOUND * use UPDATE command = TableRowID: %d  fk_CandidateUID: %d  Email %s \n' %  ( row[0], row[1], row[8]) )
        c1 = list_of_rows[i]['EmailAddress']
        c2= row[8]
        print( '   Value in Excel: %s   Value in SQL: %s' % (c1,c2))

        if  c1 != c2 : 
            print( "     >>>  EMAIL UPDATE REQUIRED  <<< \n")



print("**END**")
                   




