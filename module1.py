# from openpyxl import Workbook
from openpyxl.compat import range
# from openpyxl.cell import get_column_letter
import openpyxl as xl
import os.path
import pyodbc
import pprint

def pretty(d, indent=0):
   for key, value in d.items():
      print ( '\t' * indent + str(key))
      if isinstance(value, dict):
         pretty(value, indent+1)
      else:
         print ('\t' * (indent+1) + str(value))



print('Opening workbook...')
   
wb = xl.load_workbook('RJ_Survey0ct2013.xlsx')
print( wb.sheetnames )


ws = wb.get_sheet_by_name('Mapping')
# Mapping sheet has valid range A1:E50

print( "Case one \n")

for index, row in enumerate(ws.iter_rows('A1:E50')):
    for ccell in row:
        print(ws.cell(row=index + 1, column=1).value, ccell.value)

print("\nCase Two")

for row in ws.iter_rows('A1:E50'):    for ccell in row:        print(ccell)

sheet = wb.get_sheet_by_name('Export')


colHeads = []
list_of_rows = {}

 
print('Reading rows...')
print( "Max Rows: %d  Max Cols: %d" % (sheet.max_row, sheet.max_column) )
# print( sheet.cell(row=1,column=1).value)

# read first row ( Column Headers )
sheetrow = sheet.rows[0]

# Each row in the spreadsheet has data for one survey.
for cellObj in sheetrow:
        colHeads.append(cellObj.value)
 
# for sheetrow in rows 2 to max_rows -- READ rest of sheet
ndx = 1
for ndx in range(1, sheet.max_row):
    thisrow = sheet.rows[ndx]
    # print( hex(id(thisrow)))
    rowData = {}
    j=0
    for cellObj in thisrow :
        rowData[colHeads[j]]= cellObj.value
        j += 1   
    list_of_rows[ndx] = rowData  # add this dictionary to 'list_of_rows' Dictionary

#pretty(list_of_rows,0)


for i in range(1, 30):
     print( list_of_rows[i]['LastName'], 
           list_of_rows[i]['FirstName'], 
           list_of_rows[i]['PantherID'], 
           list_of_rows[i]['EmailAddress']
           )
